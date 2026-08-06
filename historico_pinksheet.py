# -*- coding: utf-8 -*-
"""
historico_pinksheet.py — faz o historico crescer sozinho.

PROBLEMA QUE RESOLVE
docs/historico.json e uma fotografia: 19 series mensais que terminam em jun/2026,
extraidas de um clipping. Ninguem escreve nele. Em outubro ele ainda vai terminar
em junho e o grafico do site fica com buraco na ponta.

FONTE
Pink Sheet do World Bank, arquivo CMO-Historical-Data-Monthly.xlsx. Estrutura
confirmada pelo probe de 05/08/2026 (docs/diagnostico-fontes.json, campo
pink_sheet), nao suposta:
    abas   : AFOSHEET | Monthly Prices | Monthly Indices | Description | Index Weights
    linha 4: "Updated on August 04, 2026"
    linha 5: nomes das colunas  ("Crude oil, average", "Coal, Australian", ...)
    linha 6: unidades           ("($/bbl)", "($/mt)", "($/kg)", ...)
    linha 7: dados comecam, coluna A no formato "1960M01"

O ID do documento muda todo mes ("...-0050012026"), por isso a URL NAO e fixada:
e lida da pagina indice a cada execucao, como o probe fez.

COMO O MAPEAMENTO E FEITO
Nao ha lista de nomes chumbada aqui. Cada serie do historico.json declara o proprio
benchmark ("World Bank Soybeans", "World Bank Potassium chloride"). O modulo tira o
prefixo "World Bank " e casa com o nome da coluna na linha 5. Serie cujo benchmark
nao comece com "World Bank" e ignorada de proposito: Boi gordo (CEPEA), USD/BRL
(BCB), MAP e Ureia granular (USDA) e MAP CFR Brasil (Investing.com) tem outra fonte
e nao podem ser preenchidas com dado do World Bank.

REGRAS DE GRAVACAO
  - so acrescenta mes que ainda nao existe na serie;
  - unidade do arquivo tem que ser compativel com a unidade da serie, senao a serie
    e pulada e o motivo fica na nota. Nunca converte calado;
  - valor "…" ou vazio no arquivo e ignorado, nao virou zero;
  - se um mes que ja existia voltar com valor diferente, grava o novo E registra a
    revisao na nota. Revisao do World Bank acontece; em silencio, nao.
"""
import io
import json
import re
from pathlib import Path

import requests

TIMEOUT = 120
UA = {"User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                     "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
      "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
      "Accept-Language": "en-GB,en;q=0.9,pt-BR;q=0.8"}

INDICE = "https://www.worldbank.org/en/research/commodity-markets"
ABA = "Monthly Prices"
LINHA_NOMES, LINHA_UNIDADES, LINHA_DADOS = 5, 6, 7

# Unidade do arquivo -> unidade usada no historico.json. So grafia, nunca conversao.
UNIDADES = {
    "$/mt": "US$/t", "$/dmt": "US$/t", "$/kg": "US$/kg", "$/troy oz": "US$/oz troy",
    "$/oz": "US$/oz troy", "$/dmtu": "US$/dmtu", "$/bbl": "US$/bbl",
    "$/mmbtu": "US$/mmbtu", "$/cubic meter": "US$/m3", "c/kg": "US cents/kg",
    "$/mt)": "US$/t",
}


def _limpa_nome(x):
    """Nome de coluna sem marca de nota de pe ("Coal, South African **")."""
    return re.sub(r"[*†‡]+", "", str(x or "")).strip()


def _limpa_unidade(x):
    u = str(x or "").strip().strip("()").strip().lower()
    return UNIDADES.get(u, None)


def _mes(celula):
    """"1960M01" -> "1960-01". Devolve None para qualquer outro formato."""
    m = re.fullmatch(r"\s*(\d{4})M(\d{1,2})\s*", str(celula or ""))
    if not m:
        return None
    mes = int(m.group(2))
    return f"{m.group(1)}-{mes:02d}" if 1 <= mes <= 12 else None


def _numero(x):
    """Valor do arquivo. "…", "..", vazio e texto viram None, nunca zero."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "")
    if not s or s in {"…", "...", "..", "n/a", "na", "-"}:
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    return v


def achar_xlsx(notas):
    """Le a pagina indice e devolve a URL do xlsx mensal. Nunca fixa a URL."""
    try:
        r = requests.get(INDICE, headers=UA, timeout=TIMEOUT)
        r.raise_for_status()
    except Exception as ex:
        notas.append(f"Pink Sheet: pagina indice falhou {type(ex).__name__}: {str(ex)[:90]}")
        return None
    links = re.findall(r'href="([^"]+\.xlsx?)"', r.text, re.I)
    alvo = next((l for l in links if re.search(r"Historical-Data-Monthly", l, re.I)), None)
    if not alvo:
        alvo = next((l for l in links if re.search(r"CMO|Pink", l, re.I)), None)
    if not alvo:
        notas.append(f"Pink Sheet: nenhum xlsx mensal na pagina indice "
                     f"({len(links)} xlsx vistos). Layout mudou.")
        return None
    if not alvo.startswith("http"):
        alvo = "https://www.worldbank.org" + alvo
    notas.append(f"Pink Sheet: usando {alvo[-58:]}")
    return alvo


def ler_planilha(conteudo, notas):
    """Devolve (colunas, linhas, atualizado_em).

    colunas: {nome_limpo: (indice, unidade_convertida_ou_None)}
    linhas : [(mes, [valores...]), ...]
    """
    try:
        import openpyxl
    except ImportError:
        notas.append("Pink Sheet: openpyxl ausente (pip install openpyxl)")
        return None, None, None
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    if ABA not in wb.sheetnames:
        notas.append(f"Pink Sheet: aba {ABA!r} nao existe. Abas: {wb.sheetnames}")
        return None, None, None
    ws = wb[ABA]
    grade = [list(r) for r in ws.iter_rows(values_only=True)]
    if len(grade) < LINHA_DADOS:
        notas.append(f"Pink Sheet: planilha com {len(grade)} linhas, menos que o esperado")
        return None, None, None

    atualizado = None
    for linha in grade[:4]:
        for c in linha:
            m = re.search(r"Updated on\s+(.+)", str(c or ""))
            if m:
                atualizado = m.group(1).strip()
    nomes = grade[LINHA_NOMES - 1]
    unidades = grade[LINHA_UNIDADES - 1]
    colunas = {}
    for i, n in enumerate(nomes):
        nome = _limpa_nome(n)
        if not nome or i == 0:
            continue
        colunas[nome] = (i, _limpa_unidade(unidades[i] if i < len(unidades) else None))

    linhas = []
    for linha in grade[LINHA_DADOS - 1:]:
        mes = _mes(linha[0] if linha else None)
        if mes:
            linhas.append((mes, linha))
    notas.append(f"Pink Sheet: {len(colunas)} colunas, {len(linhas)} meses, "
                 f"arquivo atualizado em {atualizado or 'data nao declarada'}")
    return colunas, linhas, atualizado


def atualizar_historico(caminho, colunas, linhas, notas):
    """Acrescenta os meses novos em cada serie cujo benchmark seja do World Bank."""
    p = Path(caminho)
    doc = json.loads(p.read_text(encoding="utf-8"))
    series = doc.get("series") or {}
    total_novos, revisoes, puladas = 0, 0, []

    for nome, s in series.items():
        bench = str(s.get("benchmark") or "")
        if not bench.lower().startswith("world bank"):
            puladas.append(f"{nome} (fonte {bench[:26]})")
            continue
        alvo = bench[len("World Bank"):].strip()
        col = colunas.get(alvo)
        if col is None:
            achou = [k for k in colunas if k.lower() == alvo.lower()]
            col = colunas.get(achou[0]) if achou else None
        if col is None:
            # sem chute: lista as colunas parecidas para o ajuste ser preciso
            chave = alvo.lower().split(",")[0].split("(")[0].strip()
            parecidas = [k for k in colunas if chave and chave in k.lower()][:5]
            notas.append(f"{nome}: coluna {alvo!r} nao existe na planilha. "
                         + (f"Colunas parecidas: {parecidas}. Ajuste o campo benchmark "
                            f"no historico.json." if parecidas
                            else "Nenhuma coluna parecida encontrada."))
            continue
        i, unidade_arq = col
        if unidade_arq and unidade_arq != s.get("unidade"):
            notas.append(f"{nome}: unidade da planilha e {unidade_arq}, a da serie e "
                         f"{s.get('unidade')}. NAO atualizada, nada foi convertido.")
            continue

        existentes = {m: v for m, v in (s.get("pontos") or [])}
        novos = 0
        for mes, linha in linhas:
            v = _numero(linha[i] if i < len(linha) else None)
            if v is None:
                continue
            if mes in existentes:
                if abs(existentes[mes] - v) > max(1e-9, abs(v) * 1e-6):
                    notas.append(f"{nome} {mes}: revisado de {existentes[mes]} para {v}, "
                                 f"valor novo gravado")
                    existentes[mes] = v
                    revisoes += 1
                continue
            existentes[mes] = v
            novos += 1
        if novos:
            s["pontos"] = sorted([[m, v] for m, v in existentes.items()], key=lambda x: x[0])
            s["de"], s["ate"] = s["pontos"][0][0], s["pontos"][-1][0]
            s["n"] = len(s["pontos"])
            total_novos += novos
            notas.append(f"{nome}: +{novos} mes(es), agora vai ate {s['ate']}")

    if puladas:
        notas.append("Series de outra fonte, nao tocadas pelo Pink Sheet: "
                     + "; ".join(puladas))
    doc["pink_sheet_atualizado_em"] = __import__("datetime").datetime.now().isoformat(timespec="seconds")
    p.write_text(json.dumps(doc, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    notas.append(f"historico.json: +{total_novos} ponto(s), {revisoes} revisao(oes)")
    return total_novos, revisoes


def rodar(caminho_historico="docs/historico.json"):
    notas = []
    url = achar_xlsx(notas)
    if not url:
        return notas
    try:
        r = requests.get(url, headers=UA, timeout=TIMEOUT)
        r.raise_for_status()
    except Exception as ex:
        notas.append(f"Pink Sheet: download falhou {type(ex).__name__}: {str(ex)[:90]}")
        return notas
    colunas, linhas, _ = ler_planilha(r.content, notas)
    if not colunas:
        return notas
    if not Path(caminho_historico).exists():
        notas.append(f"Pink Sheet: {caminho_historico} nao existe, nada a atualizar")
        return notas
    atualizar_historico(caminho_historico, colunas, linhas, notas)
    return notas


# ------------------------------------------------------------------ testes
def _planilha_falsa():
    """Reproduz a estrutura REAL vista no probe: 4 linhas de preambulo, nomes na 5,
    unidades na 6, dados da 7 em diante, coluna A no formato AAAAMmm."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ABA
    ws["A1"] = "World Bank Commodity Price Data (The Pink Sheet)"
    ws["A2"] = "monthly prices in nominal US dollars"
    ws["A3"] = "(monthly series are averages)"
    ws["A4"] = "Updated on August 04, 2026"
    for j, nome in enumerate(["", "Soybeans", "Gold", "Potassium chloride **",
                              "Iron ore, cfr spot", "Coal, Australian"], start=1):
        ws.cell(row=5, column=j, value=nome)
    for j, un in enumerate(["", "($/mt)", "($/troy oz)", "($/mt)", "($/dmtu)", "($/mt)"], start=1):
        ws.cell(row=6, column=j, value=un)
    dados = [("2026M05", 430.0, 4050.0, 351.0, 95.0, 110.0),
             ("2026M06", 435.0, 4100.0, 352.5, 94.0, 111.0),
             ("2026M07", 441.0, 4120.0, 355.0, 93.5, 112.0),
             ("2026M08", None, 4133.9, "…", 93.0, 113.0)]
    for k, linha in enumerate(dados, start=7):
        for j, v in enumerate(linha, start=1):
            ws.cell(row=k, column=j, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _testes():
    import tempfile
    notas = []
    colunas, linhas, atualizado = ler_planilha(_planilha_falsa(), notas)
    assert atualizado == "August 04, 2026", atualizado
    assert "Potassium chloride" in colunas, list(colunas)      # nota de pe removida
    assert colunas["Soybeans"][1] == "US$/t", colunas["Soybeans"]
    assert colunas["Gold"][1] == "US$/oz troy", colunas["Gold"]
    assert [m for m, _ in linhas] == ["2026-05", "2026-06", "2026-07", "2026-08"]
    print("ok  planilha:", len(colunas), "colunas,", len(linhas), "meses, atualizada em", atualizado)

    hist = {"schema": "radar-agro-historico/1", "series": {
        "Soja": {"benchmark": "World Bank Soybeans", "unidade": "US$/t", "fonte": "World Bank",
                 "pontos": [["2026-05", 430.0], ["2026-06", 435.0]], "de": "2026-05", "ate": "2026-06", "n": 2},
        "Ouro": {"benchmark": "World Bank Gold", "unidade": "US$/oz troy", "fonte": "World Bank",
                 "pontos": [["2026-06", 4100.0]], "de": "2026-06", "ate": "2026-06", "n": 1},
        "KCl": {"benchmark": "World Bank Potassium chloride", "unidade": "US$/t", "fonte": "World Bank",
                "pontos": [["2026-06", 999.0]], "de": "2026-06", "ate": "2026-06", "n": 1},
        "Minério de ferro 62% Fe": {"benchmark": "World Bank Iron ore, cfr spot",
                "unidade": "US$/t", "fonte": "World Bank",
                "pontos": [["2026-06", 94.0]], "de": "2026-06", "ate": "2026-06", "n": 1},
        "Boi gordo": {"benchmark": "CEPEA/ESALQ SP", "unidade": "R$/@", "fonte": "CEPEA",
                "pontos": [["2026-06", 320.0]], "de": "2026-06", "ate": "2026-06", "n": 1},
    }}
    with tempfile.TemporaryDirectory() as tmp:
        alvo = Path(tmp) / "historico.json"
        alvo.write_text(json.dumps(hist), encoding="utf-8")
        n2 = []
        novos, revs = atualizar_historico(alvo, colunas, linhas, n2)
        d = json.loads(alvo.read_text(encoding="utf-8"))

        soja = dict(d["series"]["Soja"]["pontos"])
        assert soja["2026-07"] == 441.0 and "2026-08" not in soja, soja   # None ignorado
        print("ok  soja: +07/2026 = 441,0 e 08/2026 ignorado (valor vazio no arquivo)")

        ouro = dict(d["series"]["Ouro"]["pontos"])
        assert ouro["2026-08"] == 4133.9 and d["series"]["Ouro"]["ate"] == "2026-08"
        print("ok  ouro: chegou a 08/2026, ate =", d["series"]["Ouro"]["ate"])

        kcl = dict(d["series"]["KCl"]["pontos"])
        assert kcl["2026-06"] == 352.5, kcl                    # revisao gravada
        assert any("revisado de 999.0 para 352.5" in x for x in n2), n2
        assert "2026-08" not in kcl                            # "…" ignorado
        print("ok  kcl: revisao 999,0 -> 352,5 registrada e '…' ignorado")

        assert d["series"]["Minério de ferro 62% Fe"]["ate"] == "2026-06"
        assert any("unidade da planilha e US$/dmtu" in x for x in n2), n2
        print("ok  minerio: unidade incompativel, serie NAO atualizada")

        assert d["series"]["Boi gordo"]["ate"] == "2026-06"
        assert any("Boi gordo" in x for x in n2)
        print("ok  boi gordo: fonte CEPEA, nao tocado pelo Pink Sheet")
        print(f"\nresumo: +{novos} pontos, {revs} revisao(oes)")
    print("todos os testes passaram")


if __name__ == "__main__":
    _testes()
