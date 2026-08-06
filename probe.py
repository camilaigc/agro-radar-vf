# -*- coding: utf-8 -*-
"""
probe.py — diagnostico de fontes. Roda no GitHub Actions (internet aberta),
NAO no sandbox do Claude. Testa varias URLs candidatas por veiculo, reporta o
que respondeu de verdade e escreve um sources.yml sugerido com as que passaram.

Nao inventa nada: se nenhuma candidata responder, o veiculo entra na lista de
falhas com o codigo HTTP e o erro exato.

Uso:  python probe.py
Saida: docs/diagnostico-fontes.json  +  tabela no log do Actions
"""
import json, sys, time
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote_plus

import re

import requests

try:
    import feedparser
except ImportError:
    feedparser = None

ROOT = Path(__file__).resolve().parent
DOCS = ROOT / "docs"
TIMEOUT, SLEEP = 25, 0.8

UA = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "application/rss+xml, application/atom+xml, application/xml;q=0.9, text/html;q=0.8, */*;q=0.5",
    "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
}

# Veiculo -> caminhos candidatos. A ordem importa: a primeira que responder vence.
CANDIDATOS = {
    "AgFeed":             ["https://agfeed.com.br/feed", "https://agfeed.com.br/feed/", "https://agfeed.com.br/rss", "https://agfeed.com.br/?feed=rss2"],
    "The AgriBiz":        ["https://theagribiz.com/feed", "https://theagribiz.com/feed/", "https://theagribiz.com/?feed=rss2", "https://theagribiz.com/rss"],
    "Noticias Agricolas": ["https://www.noticiasagricolas.com.br/rss", "https://www.noticiasagricolas.com.br/feed", "https://www.noticiasagricolas.com.br/rss/noticias", "https://noticiasagricolas.com.br/feed/"],
    "Canal Rural":        ["https://www.canalrural.com.br/feed", "https://www.canalrural.com.br/feed/", "https://www.canalrural.com.br/rss", "https://www.canalrural.com.br/?feed=rss2"],
    "Globo Rural":        ["https://globorural.globo.com/rss/ultimas/feed.xml", "https://pox.globo.com/rss/globorural", "https://globorural.globo.com/feed/"],
    "NovaCana":           ["https://www.novacana.com/rss", "https://www.novacana.com/feed", "https://www.novacana.com/rss/noticias", "https://www.novacana.com/feed/"],
    "UDOP":               ["https://www.udop.com.br/rss", "https://www.udop.com.br/feed", "https://udop.com.br/rss.xml"],
    "Brasil Mineral":     ["https://www.brasilmineral.com.br/feed", "https://www.brasilmineral.com.br/feed/", "https://brasilmineral.com.br/?feed=rss2"],
    "NeoFeed":            ["https://neofeed.com.br/feed", "https://neofeed.com.br/feed/", "https://neofeed.com.br/?feed=rss2"],
    "Brazil Journal":     ["https://braziljournal.com/feed", "https://braziljournal.com/feed/", "https://braziljournal.com/?feed=rss2"],
    "Pipeline Valor":     ["https://pipelinevalor.globo.com/rss/feed.xml", "https://pipelinevalor.globo.com/feed/"],
    "Capital Aberto":     ["https://capitalaberto.com.br/feed", "https://capitalaberto.com.br/feed/"],
    "Money Times":        ["https://www.moneytimes.com.br/feed", "https://www.moneytimes.com.br/feed/"],
    "InfoMoney":          ["https://www.infomoney.com.br/feed", "https://www.infomoney.com.br/feed/"],
    "Exame":              ["https://exame.com/feed", "https://exame.com/feed/"],
    "AgFunder News":      ["https://agfundernews.com/feed", "https://agfundernews.com/feed/"],
    "AgroPages":          ["https://news.agropages.com/rss/rsslist.xml", "https://www.agropages.com/rss/rsslist.xml", "https://www.agropages.com/feed/"],
    "Seed World":         ["https://www.seedworld.com/feed", "https://seedworld.com/feed/"],
    "SeedQuest":          ["https://www.seedquest.com/rss.xml", "https://www.seedquest.com/feed"],
    "World Grain":        ["https://www.world-grain.com/rss/topic/1-news", "https://www.world-grain.com/feed", "https://www.world-grain.com/rss"],
    "Feed Strategy":      ["https://www.feedstrategy.com/rss/topic/1-news", "https://www.feedstrategy.com/feed", "https://www.feedstrategy.com/rss"],
    "WattAgNet":          ["https://www.wattagnet.com/rss/topic/1-news", "https://www.wattagnet.com/feed", "https://www.wattagnet.com/rss"],
    "AllAboutFeed":       ["https://www.allaboutfeed.net/rss", "https://www.allaboutfeed.net/feed", "https://www.allaboutfeed.net/rss.xml"],
    "Poultry World":      ["https://www.poultryworld.net/rss", "https://www.poultryworld.net/feed", "https://www.poultryworld.net/rss.xml"],
    "Pig Progress":       ["https://www.pigprogress.net/rss", "https://www.pigprogress.net/feed", "https://www.pigprogress.net/rss.xml"],
    "The Fish Site":      ["https://thefishsite.com/rss", "https://thefishsite.com/feed", "https://www.thefishsite.com/rss"],
    "SeafoodSource":      ["https://www.seafoodsource.com/rss", "https://www.seafoodsource.com/feed"],
    "Undercurrent News":  ["https://www.undercurrentnews.com/feed", "https://www.undercurrentnews.com/feed/"],
    "PetFood Industry":   ["https://www.petfoodindustry.com/rss/topic/1-news", "https://www.petfoodindustry.com/feed", "https://www.petfoodindustry.com/rss"],
    "Mining.com":         ["https://www.mining.com/feed", "https://www.mining.com/feed/"],
    "Mining Weekly":      ["https://www.miningweekly.com/page/rss", "https://www.miningweekly.com/rss"],
    "BNamericas":         ["https://www.bnamericas.com/en/rss", "https://www.bnamericas.com/rss"],
    "CropLife":           ["https://www.croplife.com/feed", "https://www.croplife.com/feed/"],
    "Future Farming":     ["https://www.futurefarming.com/rss", "https://www.futurefarming.com/feed"],
    "Beef Central":       ["https://www.beefcentral.com/feed", "https://www.beefcentral.com/feed/"],
}

# consultas Google News: 1 por idioma so para saber se o endpoint responde daqui
GN = {
    "GoogleNews PT (teste)": "https://news.google.com/rss/search?q=" + quote_plus("aquisição agronegócio") + "&hl=pt-BR&gl=BR&ceid=BR:pt-419",
    "GoogleNews EN (teste)": "https://news.google.com/rss/search?q=" + quote_plus("agribusiness acquisition") + "&hl=en-US&gl=US&ceid=US:en",
}

# endpoints de cotacao, para saber quais estao ao alcance do Actions
COTACOES = {
    "CEPEA widget": ("https://www.cepea.org.br/br/widgetproduto.js.php?fonte=arial&tamanho=10&largura=400px"
                     "&corfundo=dbd6b2&cortexto=333333&corlinha=ede7bf"
                     "&id_indicador%5B%5D=2&id_indicador%5B%5D=77&id_indicador%5B%5D=12&id_indicador%5B%5D=178"),
    "CEPEA widget (sem www)": ("https://cepea.org.br/br/widgetproduto.js.php?fonte=arial&tamanho=10&largura=400px"
                               "&corfundo=dbd6b2&cortexto=333333&corlinha=ede7bf"
                               "&id_indicador%5B%5D=2&id_indicador%5B%5D=77&id_indicador%5B%5D=12&id_indicador%5B%5D=178"),
    "USDA AMS (CBOT)": "https://www.ams.usda.gov/mnreports/ams_2850.pdf",
    "BCB PTAX": "https://api.bcb.gov.br/dados/serie/bcdata.sgs.1/dados/ultimos/5?formato=json",
    "LBMA precos": "https://www.lbma.org.uk/prices-and-data/lbma-precious-metal-prices",
    "LME referencia": "https://www.lme.com/en/market-data/lme-reference-prices",
    "TradingEconomics cobre": "https://tradingeconomics.com/commodity/copper",
    "TradingEconomics minerio": "https://tradingeconomics.com/commodity/iron-ore",
    "CME ureia UFB": "https://www.cmegroup.com/markets/agriculture/fertilizer/urea-granular-cfr-brazil.html",
    "CME MAP MFC": "https://www.cmegroup.com/markets/agriculture/fertilizer/map-cfr-brazil.html",
    "World Bank Pink Sheet (indice)": "https://www.worldbank.org/en/research/commodity-markets",
    "World Bank docs (Pink Sheet)": "https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012025/related/CMO-Historical-Data-Monthly.xlsx",
    "USDA ERS fertilizante": "https://www.ers.usda.gov/data-products/fertilizer-use-and-price",
    "IndexMundi ureia": "https://www.indexmundi.com/commodities/?commodity=urea&months=60",
    "IndexMundi KCl": "https://www.indexmundi.com/commodities/?commodity=potassium-chloride&months=60",
    "IndexMundi DAP": "https://www.indexmundi.com/commodities/?commodity=dap-fertilizer&months=60",
    "IndexMundi TSP": "https://www.indexmundi.com/commodities/?commodity=triple-superphosphate&months=60",
    "LBMA ouro PM (json)": "https://prices.lbma.org.uk/json/gold_pm.json",
    "LBMA ouro AM (json)": "https://prices.lbma.org.uk/json/gold_am.json",
    "LBMA prata (json)": "https://prices.lbma.org.uk/json/silver.json",
    "BCB PTAX intervalo": "https://api.bcb.gov.br/dados/serie/bcdata.sgs.1/dados?formato=json&dataInicial=01/07/2026&dataFinal=05/08/2026",
    "Comex Stat docs": "https://api-comexstat.mdic.gov.br/docs",
    "Stooq ouro (csv)": "https://stooq.com/q/l/?s=xauusd&f=sd2t2ohlc&h&e=csv",
    "Stooq cobre (csv)": "https://stooq.com/q/l/?s=hg.f&f=sd2t2ohlc&h&e=csv",
    "Stooq soja CBOT (csv)": "https://stooq.com/q/l/?s=zs.f&f=sd2t2ohlc&h&e=csv",
    "Stooq milho CBOT (csv)": "https://stooq.com/q/l/?s=zc.f&f=sd2t2ohlc&h&e=csv",
    "Stooq trigo CBOT (csv)": "https://stooq.com/q/l/?s=zw.f&f=sd2t2ohlc&h&e=csv",
}

# paginas das quais queremos AMOSTRA do conteudo, para escrever o parser em cima
# do HTML real em vez de adivinhar. Guarda so trechos curtos ao redor das palavras-chave.
# amostras que exigem tratamento especial (PDF)
AMOSTRAS_PDF = {
    "USDA AMS graos": ("https://www.ams.usda.gov/mnreports/ams_2850.pdf",
                       ["soybean", "corn", "wheat", "settle"]),
}

AMOSTRAS = {
    "TE ouro":          ("https://tradingeconomics.com/commodity/gold", ["gold", "actual", "usd", "last updated"]),
    "LBMA ouro":        ("https://www.lbma.org.uk/prices-and-data/lbma-precious-metal-prices", ["gold", "usd", "am", "pm"]),
    "TE cobre":         ("https://tradingeconomics.com/commodity/copper", ["copper", "last", "usd/lb"]),
    "TE minerio":       ("https://tradingeconomics.com/commodity/iron-ore", ["iron ore", "last", "usd/t"]),
    "Stooq ouro":       ("https://stooq.com/q/l/?s=xauusd&f=sd2t2ohlc&h&e=csv", None),
    "Stooq cobre":      ("https://stooq.com/q/l/?s=hg.f&f=sd2t2ohlc&h&e=csv", None),
    "Stooq soja":       ("https://stooq.com/q/l/?s=zs.f&f=sd2t2ohlc&h&e=csv", None),
    "IndexMundi ureia": ("https://www.indexmundi.com/commodities/?commodity=urea", ["urea", "usd", "metric ton"]),
    "IndexMundi KCl":   ("https://www.indexmundi.com/commodities/?commodity=potassium-chloride", ["potassium", "usd", "metric ton"]),
}


def coletar_amostra(url, palavras):
    """Baixa a pagina e devolve trechos curtos ao redor das palavras-chave, mais os
    numeros candidatos. Nao interpreta nada: e material para escrever o parser."""
    rec = {"url": url}
    try:
        r = requests.get(url, headers=UA, timeout=TIMEOUT, allow_redirects=True)
        rec["http"] = r.status_code
        rec["bytes"] = len(r.content)
        if r.status_code != 200:
            return rec
        txt = r.text
        if palavras is None:
            rec["conteudo_bruto"] = txt[:600]
            return rec
        import re as _re
        limpo = _re.sub(r"<script[^>]*>.*?</script>", " ", txt, flags=_re.S | _re.I)
        limpo = _re.sub(r"<style[^>]*>.*?</style>", " ", limpo, flags=_re.S | _re.I)
        limpo = _re.sub(r"<[^>]+>", " ", limpo)
        limpo = _re.sub(r"\s+", " ", limpo)
        trechos = []
        low = limpo.lower()
        for p in palavras:
            i = 0
            achou = 0
            while achou < 2:
                i = low.find(p, i)
                if i < 0:
                    break
                trechos.append(limpo[max(0, i - 110): i + 170])
                i += len(p)
                achou += 1
        rec["trechos"] = trechos[:8]
        # blocos json embutidos com "last"/"price", uteis no Trading Economics
        blobs = _re.findall(r'["\'](?:last|price|close|value)["\']\s*:\s*"?(-?\d+(?:[.,]\d+)?)"?', txt)
        rec["numeros_em_json"] = blobs[:15]
    except Exception as ex:
        rec["erro"] = f"{type(ex).__name__}: {ex}"[:180]
    return rec


def coletar_amostra_pdf(url, palavras):
    """Baixa o PDF, extrai o texto e devolve as LINHAS que contem cada palavra-chave,
    com as linhas vizinhas. Sem interpretar: material bruto para escrever o parser."""
    rec = {"url": url}
    try:
        r = requests.get(url, headers=UA, timeout=60)
        rec["http"] = r.status_code
        rec["bytes"] = len(r.content)
        if r.status_code != 200:
            return rec
        try:
            from pdfminer.high_level import extract_text
        except ImportError:
            rec["erro"] = "pdfminer.six ausente no requirements.txt"
            return rec
        import io
        texto = extract_text(io.BytesIO(r.content))
        linhas = [l.rstrip() for l in texto.split("\n")]
        rec["total_linhas"] = len(linhas)
        rec["primeiras_linhas"] = [l for l in linhas[:40] if l.strip()][:20]
        achados = []
        for idx, l in enumerate(linhas):
            low = l.lower()
            if any(p in low for p in palavras):
                bloco = [x for x in linhas[max(0, idx - 1): idx + 3]]
                achados.append({"linha": idx, "conteudo": l[:200],
                                "vizinhas": [x[:200] for x in bloco]})
            if len(achados) >= 25:
                break
        rec["linhas_com_palavra"] = achados
        # datas no formato MM/DD/AAAA, para conferir a data do relatorio
        rec["datas_encontradas"] = re.findall(r"\b\d{1,2}/\d{1,2}/\d{4}\b", texto)[:10]
    except Exception as ex:
        rec["erro"] = f"{type(ex).__name__}: {ex}"[:180]
    return rec


def newest_entry_date(parsed):
    best = None
    for e in parsed.entries[:40]:
        for k in ("published_parsed", "updated_parsed"):
            st = e.get(k)
            if st:
                try:
                    d = datetime(*st[:6], tzinfo=timezone.utc)
                except Exception:
                    continue
                if best is None or d > best:
                    best = d
    return best.isoformat() if best else None


def probe(url, parse_feed=True):
    rec = {"url": url}
    t0 = time.time()
    try:
        r = requests.get(url, headers=UA, timeout=TIMEOUT, allow_redirects=True)
        rec["http"] = r.status_code
        rec["bytes"] = len(r.content)
        rec["content_type"] = (r.headers.get("Content-Type") or "")[:60]
        if r.url != url:
            rec["redirecionou_para"] = r.url
        if parse_feed and r.status_code == 200 and feedparser is not None:
            p = feedparser.parse(r.content)
            rec["itens"] = len(p.entries)
            rec["xml_malformado"] = bool(p.bozo)
            if p.bozo:
                rec["bozo_motivo"] = str(getattr(p, "bozo_exception", ""))[:120]
            rec["item_mais_recente"] = newest_entry_date(p)
            rec["ok"] = len(p.entries) > 0
        else:
            rec["ok"] = r.status_code == 200
    except Exception as ex:
        rec["ok"] = False
        rec["erro"] = f"{type(ex).__name__}: {ex}"[:180]
    rec["ms"] = int((time.time() - t0) * 1000)
    return rec


def main():
    if feedparser is None:
        print("AVISO: feedparser ausente, os feeds nao serao parseados")
    agora = datetime.now(timezone.utc).isoformat()
    resultado = {"gerado_em_utc": agora, "veiculos": {}, "google_news": {}, "cotacoes": {}}

    print("\n=== FEEDS RSS (varias candidatas por veiculo) ===")
    aprovados = {}
    for nome, urls in CANDIDATOS.items():
        tentativas = []
        vencedora = None
        for u in urls:
            rec = probe(u)
            tentativas.append(rec)
            time.sleep(SLEEP)
            if rec.get("ok"):
                vencedora = rec
                break
        resultado["veiculos"][nome] = {"vencedora": vencedora, "tentativas": tentativas}
        if vencedora:
            aprovados[nome] = vencedora["url"]
            print(f"  OK    {nome:22} {vencedora['itens']:3} itens  {vencedora['url']}")
        else:
            piores = "; ".join(
                f"{t['url'].split('//')[-1][:38]} -> " + (str(t.get('http')) if t.get('http') else t.get('erro', '?')[:40])
                for t in tentativas)
            print(f"  FALHA {nome:22} {piores}")

    print("\n=== GOOGLE NEWS ===")
    for nome, u in GN.items():
        rec = probe(u); time.sleep(SLEEP)
        resultado["google_news"][nome] = rec
        print(f"  {'OK   ' if rec.get('ok') else 'FALHA'} {nome:24} http={rec.get('http')} itens={rec.get('itens')} {rec.get('erro','')}")

    print("\n=== COTACOES ===")
    for nome, u in COTACOES.items():
        rec = probe(u, parse_feed=False); time.sleep(SLEEP)
        resultado["cotacoes"][nome] = rec
        print(f"  {'OK   ' if rec.get('ok') else 'FALHA'} {nome:26} http={rec.get('http')} bytes={rec.get('bytes')} {rec.get('erro','')}")

    print(f"\nRESUMO: {len(aprovados)}/{len(CANDIDATOS)} veiculos com feed vivo")
    print("\n--- COLE ESTE BLOCO NO sources.yml (feeds) ---")
    print("feeds:")
    for nome, u in aprovados.items():
        print(f'  - {{name: "{nome}", url: "{u}", trusted: true}}')
    print("--- fim do bloco ---")

    print("\n=== AMOSTRAS DE COTACAO (material para escrever os parsers) ===")
    resultado["amostras"] = {}
    for nome, (u, palavras) in AMOSTRAS.items():
        rec = coletar_amostra(u, palavras); time.sleep(SLEEP)
        resultado["amostras"][nome] = rec
        n = len(rec.get("trechos", []) or []) + (1 if rec.get("conteudo_bruto") else 0)
        print(f"  {nome:20} http={rec.get('http')} trechos={n} numeros_json={len(rec.get('numeros_em_json',[]) or [])} {rec.get('erro','')}")

    print("\n=== AMOSTRAS DE PDF ===")
    resultado["amostras_pdf"] = {}
    for nome, (u, palavras) in AMOSTRAS_PDF.items():
        rec = coletar_amostra_pdf(u, palavras); time.sleep(SLEEP)
        resultado["amostras_pdf"][nome] = rec
        print(f"  {nome:20} http={rec.get('http')} linhas={rec.get('total_linhas')} "
              f"com_palavra={len(rec.get('linhas_com_palavra',[]) or [])} datas={rec.get('datas_encontradas',[])[:3]} {rec.get('erro','')}")

    resultado["sugestao_feeds"] = aprovados
    DOCS.mkdir(parents=True, exist_ok=True)
    out = DOCS / "diagnostico-fontes.json"
    # Pink Sheet do World Bank. Precisa entrar AQUI, no dicionario que e gravado,
    # e nao so no print do final: a versao anterior imprimia no log do Actions e o
    # resultado ficava inacessivel para quem le o JSON. O padrao do projeto e
    # "grava no JSON e alguem le".
    print("\n=== PINK SHEET DO WORLD BANK ===")
    try:
        resultado["pink_sheet"] = probe_pink_sheet()
        print(json.dumps(resultado["pink_sheet"], ensure_ascii=False, indent=1)[:1600])
    except Exception as ex:
        resultado["pink_sheet"] = {"erro": f"{type(ex).__name__}: {str(ex)[:200]}"}
        print("  falhou:", resultado["pink_sheet"]["erro"])

    print("\n=== COMEX STAT (POST) ===")
    try:
        resultado["comexstat_post"] = probe_comexstat()
        print(" ", resultado["comexstat_post"])
    except Exception as ex:
        resultado["comexstat_post"] = {"erro": f"{type(ex).__name__}: {str(ex)[:160]}"}

    print("\n=== CEPEA VIA NAVEGADOR ===")
    try:
        resultado["cepea_navegador"] = probe_cepea_navegador()
        print(" ", resultado["cepea_navegador"])
    except Exception as ex:
        resultado["cepea_navegador"] = {"erro": f"{type(ex).__name__}: {str(ex)[:160]}"}

    out.write_text(json.dumps(resultado, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\nescrito: {out}")


# --- Teste POST do Comex Stat -------------------------------------------------
# O endpoint /general so aceita POST, entao o GET generico acima nao serve.
# Se voltar linhas=0, o filtro esta errado: testar filter "ncm" com o NCM de 8
# digitos em vez de "sh6", e conferir os nomes validos em /general/details.
def probe_comexstat():
    import json as _json
    payload = {"flow": "import", "monthDetail": True,
               "period": {"from": "2026-01", "to": "2026-05"},
               # "sh6" e recusado com {"code":400,"message":"Filtro invalido"}. A
               # variante que funciona, confirmada em producao, e "ncm" com o NCM de
               # 8 digitos como TEXTO. O probe antigo testava a errada e reportava
               # falha numa coisa que funciona.
               "filters": [{"filter": "ncm", "values": ["31054000"]}],
               "details": ["ncm"], "metrics": ["metricFOB", "metricKG"]}
    try:
        r = requests.post("https://api-comexstat.mdic.gov.br/general", json=payload,
                          headers={**UA, "Content-Type": "application/json"}, timeout=40)
        linhas = len((r.json().get("data") or {}).get("list") or []) if r.ok else 0
        amostra = _json.dumps(r.json())[:400] if r.ok else r.text[:200]
        return {"http": r.status_code, "linhas": linhas, "amostra": amostra}
    except Exception as ex:
        return {"http": None, "linhas": 0, "erro": f"{type(ex).__name__}: {str(ex)[:120]}"}


# --- Teste do CEPEA via navegador --------------------------------------------
def probe_cepea_navegador():
    try:
        from cepea_render import render_cepea
    except ImportError as ex:
        return {"ok": False, "erro": f"import: {ex}"}
    html, erro = render_cepea()
    if erro:
        return {"ok": False, "erro": erro}
    try:
        from build import parse_cepea_widget
        achados = parse_cepea_widget(html)
    except Exception as ex:
        return {"ok": False, "bytes": len(html), "erro": f"parser: {type(ex).__name__}: {ex}"}
    return {"ok": bool(achados), "bytes": len(html), "indicadores": len(achados),
            "amostra": [f"{a['label']} {a['preco_texto']} {a['unidade']} {a['data_referencia']}"
                        for a in achados]}




# --- Pink Sheet do World Bank: descobrir a estrutura antes de escrever o parser ---
# O historico.json e estatico e termina em jun/2026. Para crescer sozinho, a fonte
# das 14 series do World Bank e o arquivo mensal do Pink Sheet. O ID do documento
# muda todo mes, entao nao da para fixar URL: tem que ler da pagina indice.
# Este probe NAO grava nada. So responde tres perguntas que eu nao consigo
# responder do meu lado, porque nao alcanco worldbank.org:
#   1. a pagina indice expoe o link do xlsx mensal?
#   2. o xlsx abre com openpyxl e quais sao as abas?
#   3. em qual linha estao os cabecalhos e como os nomes das colunas aparecem?
INDICE_WB = "https://www.worldbank.org/en/research/commodity-markets"


def probe_pink_sheet():
    import io
    out = {}
    try:
        r = requests.get(INDICE_WB, headers=UA, timeout=60)
        out["indice_http"] = r.status_code
        links = re.findall(r'href="([^"]+\.xlsx?)"', r.text, re.I)
        cand = [l for l in links if re.search(r"CMO|Pink|Historical|Monthly", l, re.I)]
        out["xlsx_encontrados"] = links[:8]
        out["candidatos"] = cand[:5]
        if not cand:
            out["conclusao"] = ("nenhum link .xlsx com CMO/Pink/Historical/Monthly na pagina "
                                "indice. Precisa de outro caminho para achar o arquivo.")
            return out
        url = cand[0] if cand[0].startswith("http") else ("https://www.worldbank.org" + cand[0])
        out["url_usada"] = url
        rx = requests.get(url, headers=UA, timeout=120)
        out["xlsx_http"] = rx.status_code
        out["xlsx_bytes"] = len(rx.content)
        if not rx.ok:
            return out
        try:
            import openpyxl
        except ImportError:
            out["conclusao"] = "openpyxl nao instalado neste ambiente"
            return out
        wb = openpyxl.load_workbook(io.BytesIO(rx.content), read_only=True, data_only=True)
        out["abas"] = wb.sheetnames
        # procura a aba mensal e mostra as 8 primeiras linhas, truncadas
        nome = next((n for n in wb.sheetnames if re.search(r"month", n, re.I)), wb.sheetnames[0])
        out["aba_lida"] = nome
        ws = wb[nome]
        amostra = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, max_col=14, values_only=True)):
            amostra.append([str(c)[:22] if c is not None else None for c in row])
        out["primeiras_linhas"] = amostra
        out["conclusao"] = ("ok. Com estas linhas eu escrevo o parser localizando a coluna pelo "
                            "NOME, nao pela posicao.")
        return out
    except Exception as ex:
        out["erro"] = f"{type(ex).__name__}: {str(ex)[:200]}"
        return out


if __name__ == "__main__":
    # Tudo roda dentro de main(), inclusive Pink Sheet, Comex Stat e CEPEA, e tudo
    # e gravado em docs/diagnostico-fontes.json. Nao existe mais resultado que viva
    # so no log do Actions.
    main()
